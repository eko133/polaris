# -*- coding: utf-8 -*-
"""
Created on Thu Jun 29 16:04:23 2017

@author: samuel
"""

# 整合FT-ICR MS的python计算

# 导入所需模块
import pandas as pd
import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# 定义各个操作的函数
# 绘制气泡图
def bubble_plot(data):
    # 读取所有可能的化合物类型
    species = data['class']
    species = species.drop_duplicates()
    # 创建绘图文件夹
    if not os.path.isdir("bubble_plot"):
        os.mkdir("bubble_plot")    
    # 依次对化合物类型进行画图
    for specie in species:
        data_specie = data[data['class'] == specie]
        plt.figure()
        #自定义图片样式
        font = {'family' : 'serif',  
        'color'  : 'black',  
        'weight' : 'normal',  
        'size'   : 14,  
        } 
        plt.axis([0,60,0,16])
        plt.xlabel("Carbon Number",fontdict=font)
        plt.ylabel("DBE",fontdict=font)
        plt.text(1,14,s=specie,fontdict=font)
        #此处缩放倍数按需调整
        plt.scatter(data_specie['C'],data_specie['DBE'],s=100000*data_specie['normalized'],alpha=0.8,edgecolors='white')
        plt.savefig("bubble_plot\\%s.png"%(specie),dpi=600)
# 计算各化合物类的相对含量
def relative_abundance(data):
    # 读取所有可能的化合物类型
    species = data['class']
    species = species.drop_duplicates()
    # 创建一个空的数据表
    abundance = pd.DataFrame().astype(float)
    for specie in species:
        data_specie = data[data['class'] == specie]
        abundance.loc[specie,excel_name] = data_specie['normalized'].sum()
    # 保存到excel(https://stackoverflow.com/questions/42370977/how-to-save-a-new-sheet-in-an-existing-excel-file-using-pandas)
    book = load_workbook('relative_abundance.xlsx')
    writer = pd.ExcelWriter('relative_abundance.xlsx', engine = 'openpyxl')
    writer.book = book
    abundance.to_excel(writer, 'relative abundance')
    writer.save()
# 计算指定化合物的DBE分布
def dbe_abundance(data):
    specie = input("please specify the specie that you want to calculate: ")
    data_specie = data[data['class'] == specie]
    DBE_count = 0
    # 创建一个空的数据表
    dbeabundance = pd.DataFrame().astype(float)
    while DBE_count<21:
        data_speciedbe = data_specie[data_specie['DBE'] == DBE_count]
        dbeabundance.loc[DBE_count,excel_name] = data_speciedbe['normalized'].sum()
        DBE_count += 1
    # 保存到excel
    book = load_workbook('relative_abundance.xlsx')
    writer = pd.ExcelWriter('relative_abundance.xlsx', engine = 'openpyxl')
    writer.book = book
    dbeabundance.to_excel(writer, 'dbe abundance of %s specie'%(specie))
    writer.save()
# 计算指定化合物指定DBE的碳数分布
def carbon_abundance(data):
    specie = input("please specify the specie that you want to calculate: ")
    dbe_count = input("please specify the dbe that you want to calculate: ")
    dbe_count = int(dbe_count)
    data_cn = data[data['class'] == specie]
    data_cn = data_cn[data_cn['DBE'] == dbe_count]
    cn_count = 10
    # 创建一个空的数据表
    cnabundance = pd.DataFrame().astype(float)
    while cn_count<60:
        cnabundance.loc[cn_count,excel_name] = data_cn[data.C == cn_count]['normalized'].sum()
        cn_count += 1
    # 保存到excel
    book = load_workbook('relative_abundance.xlsx')
    writer = pd.ExcelWriter('relative_abundance.xlsx', engine = 'openpyxl')
    writer.book = book
    cnabundance.to_excel(writer, 'carbon abundance of %s,%s'%(specie,dbe_count))
    writer.save()
# 输入并切换至指定工作目录
data_dir = input("please specify the working directory: ")
os.chdir(data_dir)
# 创建数据excel
if not os.path.isfile('relative_abundance.xlsx'):
    writer = pd.ExcelWriter('relative_abundance.xlsx', engine = 'xlsxwriter')
    writer.save()
# 遍历并读取工作路径下的所有excel文件
for root, dirs, files in os.walk(data_dir):
    for excel in files:
        # 判断是否是excel
        # 将文件名分解为文件名+扩展名
        if (os.path.splitext(excel)[1] == '.xlsx') & (os.path.splitext(excel)[0] != 'relative_abundance'):
            data = pd.read_excel(excel)
            excel_name = os.path.splitext(excel)[0]
            # 筛选excel表中的数据并归一化
            data['intensity'] = data['intensity'].astype(float)
            data = data[(data.ppm>-2.0) & (data.ppm<2.0)]
            data['normalized'] = data['intensity']/data['intensity'].sum()
            # 选择所需进行的操作
            operation = input('''data successfully loaded, now specify the operation:
1. drawing the bubble plots;
2. calculate the relative abundance of different compound classes;
3. calculate the relative abundance of specie with different DBE;
4. calculate the relative abundance of the specific specie with different carbon numbers
''')
            if operation == '1':
                bubble_plot(data)
            elif operation == '2':
                relative_abundance(data)
            elif operation == '3':
                dbe_abundance(data)
            elif operation == '4':
                carbon_abundance(data)
